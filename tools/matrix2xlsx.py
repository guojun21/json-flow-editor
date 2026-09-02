#!/usr/bin/env python3
"""两个矩阵(责任矩阵 / 访问权限矩阵)直接产成 XLSX:格子样式(填充/字色/加粗/边框/列宽/冻结首行首列)照编辑器里的配色。
单源仍是 usecases.py 与 convert_access.py 里的 M 表;这里只负责「表格 → Excel」。同时写 sidecar 元数据 data/<id>.xlsx.json(标题/日期/排序)。"""
import json, os, sys, importlib.util
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.expanduser('~/PUDUAIINFRA/10_Projects/P2606_数采工厂基础设施/01_方案/用例图_排产采集'))
import usecases as U
def load(name):
    spec = importlib.util.spec_from_file_location(name, os.path.join(ROOT, "tools", name + ".py")); m = importlib.util.module_from_spec(spec)
    src = open(spec.origin, encoding="utf-8").read().split("CW, CH, RW, X0, Y0, HH")[0]   # 只要数据定义,不要画图部分
    exec(compile(src, spec.origin, "exec"), m.__dict__); return m
thin = Side(style="thin", color="C9D1D8"); BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
def fill(hexc): return PatternFill("solid", start_color=hexc.lstrip("#").upper(), end_color=hexc.lstrip("#").upper())
def put(ws, r, c, text, bg="#FFFFFF", color="#17212D", bold=False, size=11, align="center"):
    cell = ws.cell(row=r, column=c, value=text)
    cell.fill = fill(bg); cell.font = Font(name="Microsoft YaHei", size=size, bold=bold, color=color.lstrip("#").upper())
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True); cell.border = BORDER
    return cell
def finish(ws, title, legend, ncol, widths):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol); put(ws, 1, 1, title, bold=True, size=16, align="left"); ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol); put(ws, 2, 1, legend, color="#48586A", size=10, align="left"); ws.row_dimensions[2].height = 36
    for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "B4"
def responsibility():
    m = load("convert_matrix")
    wb = Workbook(); ws = wb.active; ws.title = "责任矩阵"
    roles, nodes_ = m.roles, m.nodes_
    put(ws, 3, 1, "角色 \\ 节点", bg="#17212D", color="#FFFFFF", bold=True); ws.row_dimensions[3].height = 40
    for j, n in enumerate(nodes_): put(ws, 3, j + 2, f"{j+1}. {n}", bg="#17212D", color="#FFFFFF", bold=True, size=10)
    for i, r in enumerate(roles):
        row = 4 + i; sysr = U.ACTORS[r].startswith("【"); ws.row_dimensions[row].height = 26
        put(ws, row, 1, ("«service» " if sysr else "") + r, bg="#EDF1F4" if sysr else "#FFFFFF", align="left", bold=not sysr)
        for j, n in enumerate(nodes_):
            if (r, n) in m.main: put(ws, row, j + 2, f"R {m.cnt_main[(r, n)]}/{m.total[n]}", bg="#1E755D", color="#FFFFFF", bold=True)
            elif (r, n) in m.co: put(ws, row, j + 2, f"C {m.cnt_co[(r, n)]}/{m.total[n]}", bg="#E3F1EC")
            else: put(ws, row, j + 2, "—", color="#9AA4AE")
    finish(ws, "责任矩阵（R 执行 / C 协作） · 角色 × 排产采集节点（一期）",
           "这是责任口径不是权限口径:深色 R = 执行/主责(括号 = 该角色在此节点主责的用例数 / 节点用例总数);浅色 C = 协作;「—」= 不参与。真正的访问权限见「访问权限矩阵」。上半人员角色(RBAC),下半系统/服务身份(ACL)",
           len(nodes_) + 1, [30] + [18] * len(nodes_))
    return wb, "matrix_paichan", "责任矩阵·角色×节点（09-02）", 40
def access():
    m = load("convert_access"); V, O, A = m.V, m.O, m.A
    wb = Workbook(); ws = wb.active; ws.title = "访问权限"
    put(ws, 3, 1, "角色 \\ 功能", bg="#17212D", color="#FFFFFF", bold=True); ws.row_dimensions[3].height = 40
    for j, f in enumerate(m.FUN): put(ws, 3, j + 2, f, bg="#17212D", color="#FFFFFF", bold=True, size=10)
    for i, r in enumerate(m.ROLES):
        row = 4 + i; ws.row_dimensions[row].height = 26; put(ws, row, 1, r, align="left", bold=True)
        for j, f in enumerate(m.FUN):
            perm = m.M.get(r, {}).get(f, set()); txt = " ".join(t for t, k in (("查", V), ("操", O), ("审", A)) if k in perm) or "—"
            bg = "#1E755D" if A in perm else "#5AA085" if O in perm else "#E3F1EC" if V in perm else "#FFFFFF"
            color = "#FFFFFF" if (A in perm or O in perm) else "#17212D" if V in perm else "#9AA4AE"
            put(ws, row, j + 2, txt, bg=bg, color=color, bold=A in perm)
    row = 4 + len(m.ROLES) + 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(m.FUN) + 1); put(ws, row, 1, "系统 / 服务身份（ACL，不走人员权限）", bold=True, size=12, align="left"); ws.row_dimensions[row].height = 26
    for i, (svc, desc) in enumerate(m.SVC):
        r2 = row + 1 + i; ws.row_dimensions[r2].height = 26
        put(ws, r2, 1, "«service» " + svc, bg="#EDF1F4", align="left")
        ws.merge_cells(start_row=r2, start_column=2, end_row=r2, end_column=len(m.FUN) + 1); put(ws, r2, 2, desc, align="left")
    finish(ws, "访问权限矩阵（提案） · 人员角色 × 一期功能",
           "查 = 查看;操 = 发起/修改/执行;审 = 审批/终审/解除隔离;— = 无权限。职责分离:需求方不终审、物流只做交接确认、终审限质检组长与数据运营。系统/服务身份走 ACL 单列。",
           len(m.FUN) + 1, [26] + [16] * len(m.FUN))
    return wb, "access_paichan", "访问权限矩阵·角色×功能（09-02）", 41
for fn in (responsibility, access):
    wb, fid, title, order = fn()
    out = os.path.join(ROOT, "data", fid + ".xlsx"); wb.save(out)
    json.dump({"title": title, "date": "2026-09-02", "order": order}, open(out + ".json", "w", encoding="utf-8"), ensure_ascii=False)
    print("写出", out, os.path.getsize(out), "bytes")
